import simpy
import datetime
import time
import random
import os
import pandas
import openpyxl as opxl
import win32com.client as win32
import numpy as np
from openpyxl.utils.dataframe import dataframe_to_rows

class Simulation:

    def __init__(self,serveTime,nServers,nEmployees,lunchTime):
        self.startTime = time.time()
        self.serveTime = int(serveTime)
        self.nServers = int(nServers)
        self.nEmployees = int(nEmployees)
        self.lunchTime = lunchTime
        self.data = {"Arrival Time":[],"Queue Time":[],"Inter-Arrival":[]}

    def customer(self,env, server, TBA, serveTime):

        # Simulate a period of time before the arrival of the next customer
        yield env.timeout(TBA)

        # Record time of arrival
        arriveTime = env.now
        self.data["Arrival Time"].append(arriveTime)

        # Join the queue
        with server.request() as req:
            # Wait if resource is being used
            yield req

            # Get served
            queueTime = (env.now - arriveTime)
            self.data["Queue Time"].append(queueTime)

            yield env.timeout(serveTime)

    def distribute(self,mInterval,length):
        intervals = list(random.expovariate(1.0/mInterval) for i in range(length))
        self.data["Inter-Arrival"] = intervals
        return list(np.cumsum(intervals))

    def runSimulation(self):
        # Initialise environment and resource
        env = simpy.Environment()
        cafe = simpy.Resource(env, capacity=self.nServers)

        # Generate customers
        meanTBA = int(self.lunchTime)/(int(self.nEmployees)-1)
        arrivalTimes = self.distribute(meanTBA,self.nEmployees)
        for i in range(self.nEmployees):
            env.process(self.customer(env, cafe, arrivalTimes[i], self.serveTime))

        # Run simulation
        env.run()

    def exportData(self):
        simDay = datetime.datetime.now()
        sheetName = simDay.strftime("%H %M %S %f")
        direcName = os.path.dirname(os.path.abspath(__file__)) + "\\" + fileName

        # Export data collected to Excel
        try:
            wb = opxl.load_workbook(filename=fileName)
        except FileNotFoundError:
            new_wb = opxl.Workbook()
            new_wb.save("%s.xlsx" % self.dist)
            wb = new_wb

        ws = wb.create_sheet(title="%s" % sheetName)
        df = pandas.DataFrame.from_dict(self.data)

        for row in dataframe_to_rows(df, index=True, header=True):
            ws.append(row)

        # Format table
        ws['A1'] = 'ID #'

        #Summarise data
        ws['H2'] = 'Summary'
        ws['H3'],ws['I3'] = 'Avg. Queue Time',np.mean(self.data["Queue Time"])
        ws['H4'],ws['I4'] = '# of Servers',self.nServers
        ws['H5'],ws['I5'] = 'Serve Time (s)',self.serveTime
        ws['H6'],ws['I6'] = 'Lunch Time (hrs)',round(self.lunchTime/3600,1)
        ws['H7'],ws['I7'] = 'Mean TBA (s)',np.mean(self.data["Inter-Arrival"])
        ws['H8'],ws['I8'] = 'Spread',np.std(self.data["Queue Time"])

        # Save workbook
        wb.save(fileName)

        # Autofit columns
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        wb = excel.Workbooks.Open(direcName)
        ws = wb.Worksheets(sheetName)
        ws.Columns.AutoFit()
        wb.Save()
        excel.Application.Quit()

    def printRuntime(self):
        print('RUNTIME: --- %.2f seconds ---' % (time.time()-self.startTime))

    def run(self):
        self.runSimulation()
        self.exportData()
        self.printRuntime()

if __name__ == '__main__':
	
    startTime = time.time()
	
	#Input Data
    serveTimes = [5,10,15,30,45]
    numServers = [1,2,3,5]      
    nEmployees = 1250            
    lunchTimes = [1800,3600,7200]
	
    fileName = "FCFS.xlsx"
    for sTime in serveTimes:
        for num in numServers:
            for lunchTime in lunchTimes:
                s = Simulation(sTime,num,nEmployees,lunchTime)
                s.run()
				
    print("TOTAL RUNTIME: %.2f seconds" % (time.time() - startTime))
